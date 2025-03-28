import json
from openpyxl import load_workbook

def excel_to_json(input_excel):
    wb = load_workbook(input_excel)
    ws = wb.active
    
    result = {
        "purpose": {"title": "Experiment Purpose", "content": ""},
        "nature": {"title": "Experiment Nature", "content": ""},
        "pathogen": {"title": "Pathogen", "content": ""},
        "test_articles": {"title": "Test Articles", "content": ""},
        "positive_control": {"title": "Positive Control", "content": ""},
        "solvent_control": {"title": "Solvent Control", "content": ""},
        "experimental_sys": {"title": "Experimental System", "quantity": "", "species": "", "male": "", "female": ""},
        "methods": {"title": "Methods", "sections": []},
        "experimental_groups": {"title": "Experimental Groups", "columns": [], "rows": []},
        "detection_indicators": {"title": "Detection Indicators", "items": []}
    }
    
    current_section = None
    current_method = None
    
    for row in ws.iter_rows(values_only=True):
        if not any(row):  # 跳过空行
            continue
            
        # 解析基本信息
        if row[0] == "实验目的":
            result["purpose"]["content"] = row[1]
        elif row[0] == "实验性质":
            result["nature"]["content"] = row[1]
        elif row[0] == "病原体":
            result["pathogen"]["content"] = row[1]
        elif row[0] == "供试品":
            result["test_articles"]["content"] = row[1]
        elif row[0] == "阳性对照":
            result["positive_control"]["content"] = row[1]
        elif row[0] == "溶剂对照":
            result["solvent_control"]["content"] = row[1]
            
        # 解析实验系统
        elif row[0] == "数量":
            result["experimental_sys"]["quantity"] = row[1]
        elif row[0] == "物种":
            result["experimental_sys"]["species"] = row[1]
        elif row[0] == "雄性":
            result["experimental_sys"]["male"] = row[1]
        elif row[0] == "雌性":
            result["experimental_sys"]["female"] = row[1]
                
        # 解析检测指标
        elif row[0] == "检测指标":
            current_section = "indicators"
        elif current_section == "indicators":
            if row[0] and row[1]:
                result["detection_indicators"]["items"].append({
                    "title": row[0],
                    "content": row[1]
                })

        # 解析实验分组
        elif row[0] == "实验分组":
            current_section = "groups"
        elif current_section == "groups":
            if row[0] =="Group ID":
                result["experimental_groups"]["columns"] = list(row)
            elif row[1]: 
                result["experimental_groups"]["rows"].append(list(row))
            else:  # 遇到空行表示分组数据结束
                break
        
        # 解析实验方法
        elif row[0] == "实验方法":
            current_section = "methods"
        elif current_section == "methods":
            if row[0] and not row[1]:  # 方法标题
                current_method = {"title": row[0], "steps": [], "content": ""}
                result["methods"]["sections"].append(current_method)
            elif row[1]:  # 方法步骤
                current_method["steps"].append(row[1])
    
    return result

if __name__ == "__main__":
    input_excel = "自定义文件名.xlsx"  # 默认输入文件
    output_json = "restored_data6.json"  # 输出JSON文件
    
    try:
        data = excel_to_json(input_excel)
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"JSON文件已生成: {output_json}")
    except Exception as e:
        print(f"转换失败: {str(e)}")