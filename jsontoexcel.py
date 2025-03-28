import json
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

import json
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def load_json_data(json_file):
    with open(json_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def generate_excel(data, output_filename="experiment_details.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "实验方案汇总"
    
    # 1. 基本信息
    ws.append(["实验目的", data["purpose"]["content"]])
    ws.append(["实验性质", data["nature"]["content"]])
    ws.append(["病原体", data["pathogen"]["content"]])
    ws.append(["供试品", data["pathogen"]["content"]])
    ws.append(["阳性对照", data["positive_control"]["content"]])
    ws.append(["溶剂对照", data["solvent_control"]["content"]])
    ws.append([])  # 空行分隔
    
    # 2. 实验系统
    ws.append(["实验系统"])
    ws.append(["数量", data["experimental_sys"]["quantity"]])
    ws.append(["物种", data["experimental_sys"]["species"]])
    ws.append(["雄性", data["experimental_sys"]["male"]])
    ws.append(["雌性", data["experimental_sys"]["female"]])
    ws.append([])
    
    # 3. 实验方法
    ws.append(["实验方法"])
    for section in data["methods"]["sections"]:
        ws.append([section["title"]])
        for step in section["steps"]:
            ws.append(["", step])
        ws.append([])
    
    # 4. 实验分组
    ws.append(["实验分组"])
    ws.append(data["experimental_groups"]["columns"])
    for row in data["experimental_groups"]["rows"]:
        ws.append(row)
    ws.append([])
    
    # 5. 检测指标
    ws.append(["检测指标"])
    for item in data["detection_indicators"]["items"]:
        ws.append([item["title"], item["content"]])
    
    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_filename)
    print(f"Excel文件已生成: {output_filename}")

if __name__ == "__main__":
    # 处理命令行参数
    if len(sys.argv) > 1:
        output_file = sys.argv[1]
    else:
        output_file = "experiment_details.xlsx"
    
    try:
        # 从wwsolution.json加载数据
        data = load_json_data("wwsolution.json")
        generate_excel(data, output_file)
        print(f"文件已成功生成: {output_file}")
    except FileNotFoundError:
        print("错误: 未找到wwsolution.json文件")
    except json.JSONDecodeError:
        print("错误: wwsolution.json文件格式不正确")
    except Exception as e:
        print(f"生成文件失败: {str(e)}")